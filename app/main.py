"""Streamlit dashboard for Employee Attrition Prediction."""

import io
import os
import sys
from pathlib import Path

import joblib
import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

# Add project root to path
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.ingest import download_dataset
from src.predict import load_model, predict_batch, predict_single
from src.preprocess import (
    CATEGORICAL_FEATURES,
    NUMERIC_FEATURES,
    split_features_target,
)

# --- Constants ---
IS_CLOUD = os.environ.get("STREAMLIT_SHARING_MODE") == "streamlit_sharing"
MODEL_PATH = PROJECT_ROOT / "models" / "best_model.joblib"

COLORS = {
    "primary": "#1E3A5F",
    "high": "#E74C3C",
    "medium": "#F39C12",
    "low": "#27AE60",
}

TIER_COLORS = {"High": COLORS["high"], "Medium": COLORS["medium"], "Low": COLORS["low"]}


# --- Page config ---
st.set_page_config(
    page_title="Employee Attrition Predictor",
    page_icon="👥",
    layout="wide",
    initial_sidebar_state="expanded",
)


@st.cache_data
def get_dataset() -> pd.DataFrame:
    return download_dataset()


@st.cache_resource
def get_model():
    try:
        return load_model(MODEL_PATH)
    except FileNotFoundError:
        return None


# --- Sidebar ---
def render_sidebar():
    st.sidebar.title("👥 Attrition Predictor")
    st.sidebar.markdown("---")

    df = get_dataset()
    attrition_rate = (df["Attrition"] == "Yes").mean()

    st.sidebar.metric("Total Employees", f"{len(df):,}")
    st.sidebar.metric("Attrition Rate", f"{attrition_rate:.1%}")

    artifact = get_model()
    if artifact:
        st.sidebar.metric("Model", artifact["model_name"])
        st.sidebar.metric("Trained", artifact["trained_at"][:10])
    else:
        st.sidebar.warning("No model loaded")

    st.sidebar.markdown("---")

    page = st.sidebar.radio(
        "Navigation",
        ["Overview", "Model Performance", "Flight Risk Scorer", "Workforce Risk Report"],
    )

    if not IS_CLOUD:
        st.sidebar.markdown("---")
        if st.sidebar.button("🔄 Retrain Model", use_container_width=True):
            with st.spinner("Retraining model..."):
                from src.train import train_and_evaluate
                train_and_evaluate()
                st.cache_resource.clear()
                st.rerun()

    return page


# --- Page: Overview ---
def page_overview():
    st.title("📊 Workforce Overview")
    st.markdown("Key workforce statistics and attrition insights.")

    df = get_dataset()
    attrition_yes = df[df["Attrition"] == "Yes"]
    attrition_no = df[df["Attrition"] == "No"]

    # KPI row
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Total Employees", f"{len(df):,}")
    col2.metric("Attrition Count", f"{len(attrition_yes):,}")
    col3.metric("Attrition Rate", f"{len(attrition_yes)/len(df):.1%}")
    col4.metric("Avg Monthly Income", f"${df['MonthlyIncome'].mean():,.0f}")

    st.markdown("---")

    # Charts row 1
    col1, col2 = st.columns(2)

    with col1:
        fig = px.pie(
            df, names="Attrition", title="Attrition Distribution",
            color="Attrition",
            color_discrete_map={"Yes": COLORS["high"], "No": COLORS["primary"]},
            hole=0.4,
        )
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        dept_attrition = df.groupby("Department")["Attrition"].apply(
            lambda x: (x == "Yes").mean()
        ).reset_index()
        dept_attrition.columns = ["Department", "Attrition Rate"]
        fig = px.bar(
            dept_attrition, x="Department", y="Attrition Rate",
            title="Attrition Rate by Department",
            color="Attrition Rate",
            color_continuous_scale=[[0, COLORS["low"]], [0.5, COLORS["medium"]], [1, COLORS["high"]]],
        )
        fig.update_layout(
            yaxis_tickformat=".0%",
            margin=dict(t=40, b=0, l=0, r=0),
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig, use_container_width=True)

    # Charts row 2
    col1, col2 = st.columns(2)

    with col1:
        fig = px.histogram(
            df, x="Age", color="Attrition", barmode="overlay",
            title="Age Distribution by Attrition",
            color_discrete_map={"Yes": COLORS["high"], "No": COLORS["primary"]},
            opacity=0.7,
        )
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        ot_attrition = df.groupby("OverTime")["Attrition"].apply(
            lambda x: (x == "Yes").mean()
        ).reset_index()
        ot_attrition.columns = ["OverTime", "Attrition Rate"]
        fig = px.bar(
            ot_attrition, x="OverTime", y="Attrition Rate",
            title="Attrition Rate by Overtime Status",
            color="OverTime",
            color_discrete_map={"Yes": COLORS["high"], "No": COLORS["low"]},
        )
        fig.update_layout(
            yaxis_tickformat=".0%",
            margin=dict(t=40, b=0, l=0, r=0),
            showlegend=False,
        )
        st.plotly_chart(fig, use_container_width=True)

    # Charts row 3
    col1, col2 = st.columns(2)

    with col1:
        fig = px.box(
            df, x="Attrition", y="MonthlyIncome",
            title="Monthly Income by Attrition",
            color="Attrition",
            color_discrete_map={"Yes": COLORS["high"], "No": COLORS["primary"]},
        )
        fig.update_layout(margin=dict(t=40, b=0, l=0, r=0))
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        role_attrition = df.groupby("JobRole")["Attrition"].apply(
            lambda x: (x == "Yes").mean()
        ).reset_index()
        role_attrition.columns = ["JobRole", "Attrition Rate"]
        role_attrition = role_attrition.sort_values("Attrition Rate", ascending=True)
        fig = px.bar(
            role_attrition, x="Attrition Rate", y="JobRole",
            title="Attrition Rate by Job Role", orientation="h",
            color="Attrition Rate",
            color_continuous_scale=[[0, COLORS["low"]], [0.5, COLORS["medium"]], [1, COLORS["high"]]],
        )
        fig.update_layout(
            xaxis_tickformat=".0%",
            margin=dict(t=40, b=0, l=0, r=0),
            coloraxis_showscale=False,
        )
        st.plotly_chart(fig, use_container_width=True)


# --- Page: Model Performance ---
def page_model_performance():
    st.title("🎯 Model Performance")

    artifact = get_model()
    if artifact is None:
        st.error("Model not found. Please run `python src/train.py` and commit the model file.")
        return

    results = artifact["results"]

    # Comparison table
    st.subheader("Model Comparison")
    comparison_data = []
    for name, metrics in results.items():
        comparison_data.append({
            "Model": name,
            "ROC-AUC": f"{metrics['roc_auc']:.4f}",
            "F1 Score": f"{metrics['f1']:.4f}",
            "Precision": f"{metrics['precision']:.4f}",
            "Recall": f"{metrics['recall']:.4f}",
            "CV AUC (mean)": f"{metrics['cv_mean_auc']:.4f}",
        })
    comparison_df = pd.DataFrame(comparison_data)
    st.dataframe(comparison_df, use_container_width=True, hide_index=True)

    st.markdown(f"**Best Model:** {artifact['model_name']}")

    st.markdown("---")
    col1, col2 = st.columns(2)

    # ROC Curves
    with col1:
        st.subheader("ROC Curves")
        fig = go.Figure()
        colors = [COLORS["primary"], COLORS["medium"], COLORS["high"]]
        for i, (name, metrics) in enumerate(results.items()):
            fig.add_trace(go.Scatter(
                x=metrics["fpr"], y=metrics["tpr"],
                mode="lines",
                name=f"{name} (AUC={metrics['roc_auc']:.3f})",
                line=dict(color=colors[i % len(colors)], width=2),
            ))
        fig.add_trace(go.Scatter(
            x=[0, 1], y=[0, 1], mode="lines",
            name="Random", line=dict(dash="dash", color="gray"),
        ))
        fig.update_layout(
            xaxis_title="False Positive Rate",
            yaxis_title="True Positive Rate",
            margin=dict(t=10, b=0, l=0, r=0),
            legend=dict(x=0.4, y=0.1),
        )
        st.plotly_chart(fig, use_container_width=True)

    # Confusion Matrix (best model)
    with col2:
        st.subheader(f"Confusion Matrix — {artifact['model_name']}")
        best_metrics = results[artifact["model_name"]]
        cm = np.array(best_metrics["confusion_matrix"])
        labels = ["Stayed", "Left"]
        fig = px.imshow(
            cm, text_auto=True,
            x=labels, y=labels,
            color_continuous_scale=[[0, "#F0F2F6"], [1, COLORS["primary"]]],
            labels=dict(x="Predicted", y="Actual"),
        )
        fig.update_layout(margin=dict(t=10, b=0, l=0, r=0), coloraxis_showscale=False)
        st.plotly_chart(fig, use_container_width=True)

    # Feature importance
    st.markdown("---")
    st.subheader("Feature Importance (Top 20)")
    feature_names = artifact["feature_names"]
    classifier = artifact["pipeline"].named_steps["classifier"]

    if hasattr(classifier, "feature_importances_"):
        importances = classifier.feature_importances_
    elif hasattr(classifier, "coef_"):
        importances = np.abs(classifier.coef_).flatten()
    else:
        importances = np.zeros(len(feature_names))

    n = min(len(feature_names), len(importances))
    fi_df = pd.DataFrame({
        "Feature": feature_names[:n],
        "Importance": importances[:n],
    }).sort_values("Importance", ascending=True).tail(20)

    fig = px.bar(
        fi_df, x="Importance", y="Feature", orientation="h",
        color="Importance",
        color_continuous_scale=[[0, COLORS["low"]], [0.5, COLORS["medium"]], [1, COLORS["primary"]]],
    )
    fig.update_layout(
        margin=dict(t=10, b=0, l=0, r=0),
        coloraxis_showscale=False,
        height=500,
    )
    st.plotly_chart(fig, use_container_width=True)


# --- Page: Flight Risk Scorer ---
def page_flight_risk_scorer():
    st.title("🔍 Flight Risk Scorer")
    st.markdown("Enter individual employee attributes to predict attrition risk.")

    artifact = get_model()
    if artifact is None:
        st.error("Model not found. Please run `python src/train.py` and commit the model file.")
        return

    df = get_dataset()

    with st.form("employee_form"):
        col1, col2, col3 = st.columns(3)

        with col1:
            age = st.number_input("Age", min_value=18, max_value=65, value=35)
            monthly_income = st.number_input("Monthly Income", min_value=1000, max_value=20000, value=5000)
            distance = st.number_input("Distance From Home", min_value=1, max_value=30, value=10)
            years_company = st.number_input("Years at Company", min_value=0, max_value=40, value=5)
            years_role = st.number_input("Years in Current Role", min_value=0, max_value=18, value=3)
            total_working = st.number_input("Total Working Years", min_value=0, max_value=40, value=10)
            num_companies = st.number_input("Num Companies Worked", min_value=0, max_value=9, value=2)
            years_promotion = st.number_input("Years Since Last Promotion", min_value=0, max_value=15, value=1)

        with col2:
            department = st.selectbox("Department", sorted(df["Department"].unique()))
            job_role = st.selectbox("Job Role", sorted(df["JobRole"].unique()))
            business_travel = st.selectbox("Business Travel", sorted(df["BusinessTravel"].unique()))
            overtime = st.selectbox("OverTime", ["Yes", "No"])
            gender = st.selectbox("Gender", ["Male", "Female"])
            marital = st.selectbox("Marital Status", ["Single", "Married", "Divorced"])
            education_field = st.selectbox(
                "Education Field",
                sorted(df["EducationField"].unique()) if "EducationField" in df.columns else ["Life Sciences"],
            )

        with col3:
            job_satisfaction = st.slider("Job Satisfaction", 1, 4, 3)
            env_satisfaction = st.slider("Environment Satisfaction", 1, 4, 3)
            job_involvement = st.slider("Job Involvement", 1, 4, 3)
            work_life = st.slider("Work-Life Balance", 1, 4, 3)
            relationship = st.slider("Relationship Satisfaction", 1, 4, 3)
            education = st.slider("Education", 1, 5, 3)
            job_level = st.slider("Job Level", 1, 5, 2)
            stock_option = st.slider("Stock Option Level", 0, 3, 1)
            perf_rating = st.selectbox("Performance Rating", [3, 4])
            pct_hike = st.number_input("Percent Salary Hike", min_value=11, max_value=25, value=15)
            training = st.number_input("Training Times Last Year", min_value=0, max_value=6, value=3)
            years_manager = st.number_input("Years With Current Manager", min_value=0, max_value=17, value=3)

        submitted = st.form_submit_button("Predict Risk", use_container_width=True, type="primary")

    if submitted:
        employee = {
            "Age": age,
            "BusinessTravel": business_travel,
            "DailyRate": 800,
            "Department": department,
            "DistanceFromHome": distance,
            "Education": education,
            "EducationField": education_field,
            "EnvironmentSatisfaction": env_satisfaction,
            "Gender": gender,
            "HourlyRate": 65,
            "JobInvolvement": job_involvement,
            "JobLevel": job_level,
            "JobRole": job_role,
            "JobSatisfaction": job_satisfaction,
            "MaritalStatus": marital,
            "MonthlyIncome": monthly_income,
            "MonthlyRate": 14000,
            "NumCompaniesWorked": num_companies,
            "OverTime": overtime,
            "PercentSalaryHike": pct_hike,
            "PerformanceRating": perf_rating,
            "RelationshipSatisfaction": relationship,
            "StockOptionLevel": stock_option,
            "TotalWorkingYears": total_working,
            "TrainingTimesLastYear": training,
            "WorkLifeBalance": work_life,
            "YearsAtCompany": years_company,
            "YearsInCurrentRole": years_role,
            "YearsSinceLastPromotion": years_promotion,
            "YearsWithCurrManager": years_manager,
        }

        result = predict_single(employee, artifact)

        st.markdown("---")

        # Display results
        tier = result["risk_tier"]
        score = result["risk_score"]
        color = TIER_COLORS[tier]

        col1, col2, col3 = st.columns(3)
        col1.metric("Risk Score", f"{score}/100")
        col2.markdown(
            f"### Risk Tier: <span style='color:{color};font-weight:bold'>{tier}</span>",
            unsafe_allow_html=True,
        )
        col3.metric("Probability", f"{result['probability']:.1%}")

        # Risk gauge
        fig = go.Figure(go.Indicator(
            mode="gauge+number",
            value=score,
            title={"text": "Flight Risk Score"},
            gauge={
                "axis": {"range": [0, 100]},
                "bar": {"color": color},
                "steps": [
                    {"range": [0, 30], "color": "#E8F5E9"},
                    {"range": [30, 60], "color": "#FFF3E0"},
                    {"range": [60, 100], "color": "#FFEBEE"},
                ],
                "threshold": {
                    "line": {"color": "black", "width": 2},
                    "thickness": 0.75,
                    "value": score,
                },
            },
        ))
        fig.update_layout(height=250, margin=dict(t=40, b=0, l=40, r=40))
        st.plotly_chart(fig, use_container_width=True)

        # Top SHAP drivers
        st.subheader("Top 5 Risk Drivers")
        shap_data = result["shap_values"][:5]
        shap_df = pd.DataFrame(shap_data)

        fig = px.bar(
            shap_df, x="shap_value", y="feature", orientation="h",
            title="SHAP Feature Contributions",
            color="shap_value",
            color_continuous_scale=[[0, COLORS["low"]], [0.5, "#F0F2F6"], [1, COLORS["high"]]],
            color_continuous_midpoint=0,
        )
        fig.update_layout(
            yaxis=dict(autorange="reversed"),
            margin=dict(t=40, b=0, l=0, r=0),
            coloraxis_showscale=False,
            height=300,
        )
        st.plotly_chart(fig, use_container_width=True)


# --- Page: Workforce Risk Report ---
def page_workforce_risk_report():
    st.title("📋 Workforce Risk Report")
    st.markdown("Upload a CSV of employees to generate a ranked flight risk report.")

    artifact = get_model()
    if artifact is None:
        st.error("Model not found. Please run `python src/train.py` and commit the model file.")
        return

    uploaded = st.file_uploader("Upload employee CSV", type=["csv"])

    if uploaded is not None:
        try:
            input_df = pd.read_csv(uploaded)
            st.success(f"Loaded {len(input_df)} employees")

            with st.spinner("Scoring employees..."):
                result_df = predict_batch(input_df, artifact)

            # Summary metrics
            col1, col2, col3, col4 = st.columns(4)
            col1.metric("Total Employees", len(result_df))
            col2.metric(
                "High Risk",
                len(result_df[result_df["risk_tier"] == "High"]),
            )
            col3.metric(
                "Medium Risk",
                len(result_df[result_df["risk_tier"] == "Medium"]),
            )
            col4.metric(
                "Low Risk",
                len(result_df[result_df["risk_tier"] == "Low"]),
            )

            # Risk distribution chart
            tier_counts = result_df["risk_tier"].value_counts().reset_index()
            tier_counts.columns = ["Risk Tier", "Count"]
            fig = px.bar(
                tier_counts, x="Risk Tier", y="Count",
                color="Risk Tier",
                color_discrete_map=TIER_COLORS,
                title="Risk Tier Distribution",
            )
            fig.update_layout(
                margin=dict(t=40, b=0, l=0, r=0),
                showlegend=False,
            )
            st.plotly_chart(fig, use_container_width=True)

            # Results table
            st.subheader("Ranked Flight Risk Table")

            def color_tier(val):
                colors = {
                    "High": f"background-color: {COLORS['high']}; color: white",
                    "Medium": f"background-color: {COLORS['medium']}; color: white",
                    "Low": f"background-color: {COLORS['low']}; color: white",
                }
                return colors.get(val, "")

            display_cols = [c for c in result_df.columns if c in [
                "EmployeeNumber", "Age", "Department", "JobRole",
                "MonthlyIncome", "OverTime", "YearsAtCompany",
                "risk_score", "risk_tier",
            ]]
            if not display_cols:
                display_cols = list(result_df.columns)

            styled = result_df[display_cols].style.map(
                color_tier, subset=["risk_tier"] if "risk_tier" in display_cols else [],
            )
            st.dataframe(styled, use_container_width=True, height=400)

            # Download as Excel
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                result_df.to_excel(writer, sheet_name="Flight Risk Report", index=False)

                workbook = writer.book
                worksheet = writer.sheets["Flight Risk Report"]

                from openpyxl.styles import Font, PatternFill

                # Find the risk_tier column
                tier_col_idx = None
                for idx, col in enumerate(result_df.columns, 1):
                    if col == "risk_tier":
                        tier_col_idx = idx
                        break

                if tier_col_idx:
                    fills = {
                        "High": PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid"),
                        "Medium": PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid"),
                        "Low": PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid"),
                    }
                    white_font = Font(color="FFFFFF", bold=True)

                    for row_idx in range(2, len(result_df) + 2):
                        cell = worksheet.cell(row=row_idx, column=tier_col_idx)
                        tier_value = cell.value
                        if tier_value in fills:
                            cell.fill = fills[tier_value]
                            cell.font = white_font

            st.download_button(
                label="📥 Download Excel Report",
                data=output.getvalue(),
                file_name="flight_risk_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        except Exception as e:
            st.error(f"Error processing file: {e}")
    else:
        st.info("Upload a CSV file with employee data to generate predictions.")
        st.markdown("**Expected columns include:** Age, Department, JobRole, MonthlyIncome, OverTime, "
                    "YearsAtCompany, JobSatisfaction, EnvironmentSatisfaction, etc.")


# --- Main ---
def main():
    page = render_sidebar()

    if page == "Overview":
        page_overview()
    elif page == "Model Performance":
        page_model_performance()
    elif page == "Flight Risk Scorer":
        page_flight_risk_scorer()
    elif page == "Workforce Risk Report":
        page_workforce_risk_report()


if __name__ == "__main__":
    main()
